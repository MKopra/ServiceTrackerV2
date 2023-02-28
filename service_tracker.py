from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, range_boundaries
from openpyxl.styles import Font, Border
import csv 

#print("Enter file by dragging Excel you wish to filter into this terminal window.")

print("\nWelcome to Services Tracker, this application is made to filter large GCSSA spreadsheets down to readable size.")
print("You'll be prompted to enter your company, please enter as an abbreviation for the program to function correctly.")
print("Once you've entered your company in uppercase, you'll enter the type of date you want to be shown.")
print("The program will run for 3-4 minutes and create a second sheet on the raw data GCSSA spreadsheet you downloaded")
print("When this terminal closes, you'll be able to open the file and see the filtered second sheet. Enjoy!\n")
            

print("Enter your Company (Ex: A,B,C,HHC,FSC)\n")
userinput_co = input()
print("What type of date would you like (Ex: early, planned, or late)\n")
userinput_date = input()


wb = load_workbook("services.xlsx")
firstsheet = wb.active
outputsheet = wb.create_sheet('program output')
#outputsheet = wb.copy_worksheet(firstsheet) # new strategy of copying workbook and then deleting stuff from copy
wb.save("services.xlsx")


header_range = firstsheet['1']
output_header = outputsheet['A1':'Z1']

#print(output_header)

if userinput_date in ("early","Early","EARLY","earluy","erly", "elry"):
    for cell in header_range:
        if cell.value in ("Main work center","Completion Status","Admin No.","Model number","Description of technical object","Early Date"):
            outputsheet.cell(row=cell.row, column=cell.column, value=cell.value)
            #if outputsheet.cell.value is None:
            #    idx = outputsheet.cell.column
                #print(idx)
            #    outputsheet.delete_cols(idx, 1)
if userinput_date in ("plan","planned","Planned","Plan date","plan date","PLANNED","PLAN"):
    for cell in header_range:
        if cell.value in ("Main work center","Completion Status","Admin No.","Model number","Description of technical object","PlanDate MaintCall"):
            outputsheet.cell(row=cell.row, column=cell.column, value=cell.value)
if userinput_date in ("late","late date","LATE","Late Date","Late date"):
    for cell in header_range:
        if cell.value in ("Main work center","Completion Status","Admin No.","Model number","Description of technical object","Late Date"):
            outputsheet.cell(row=cell.row, column=cell.column, value=cell.value)

#for cell in output_header:
 #   if cell.value is None:
 #       idx = outputsheet.cell.column
 #       print(idx)
 #       outputsheet.delete_cols(idx, 1)


# captains log midnight -- idk if its best to do this by deleting columns and rows from a copy or copying copies and rows onto a blank - ask james 

wb.save("services.xlsx")
uic_column = 0 #firstsheet['A']
min_col, min_row, max_col, max_row = range_boundaries("A:M")

if userinput_co in ("a","A","A co", "Alpha"):
    for row in firstsheet.iter_rows():
        if row[uic_column].value == "WAD8A0":# "Main work center"): # or main work center?
            outputsheet.append((cell.value for cell in row[min_col-1:max_col]))
            #outputsheet.cell(row, paste_uic).value = firstsheet.cell(row, uic_column).value
                #row_id = cell.row
                #print(row_id)
                #outputsheet.delete_rows(row_id, 1)
if userinput_co in ("b","B","B co","Bravo"):
    for row in firstsheet.iter_rows():
        if row[uic_column].value == "WAD8B0":
            outputsheet.append((cell.value for cell in row[min_col-1:max_col]))
if userinput_co in ("c","C","C co","Charlie"):
    for row in firstsheet.iter_rows():
        if row[uic_column].value == "WAD8C0":
            outputsheet.append((cell.value for cell in row[min_col-1:max_col]))
if userinput_co in ("hhc","HHC","HHC co","HQ"):
    for row in firstsheet.iter_rows():
        if row[uic_column].value == "WAD8T0":
            outputsheet.append((cell.value for cell in row[min_col-1:max_col]))
if userinput_co in ("fsc","FSC","HFSC","H FSC"):
    for row in firstsheet.iter_rows():
        if row[uic_column].value == "WH0KH0":
            outputsheet.append((cell.value for cell in row[min_col-1:max_col]))
                
for cell in output_header:
    for blank in cell:    
        if blank.value is None:
            idx = blank.column
            outputsheet.delete_cols(idx, 1)

# header_style = outputsheet.row_dimensions[1]
# header_style.font = Font(bold=True)
uic_head = outputsheet['A1']
comp_head = outputsheet['B1']
admin_head = outputsheet['C1']
mod_head = outputsheet['D1']
desc_head = outputsheet['E1']
date_head = outputsheet['F1']
uic_head.font = Font(bold=True)
comp_head.font = Font(bold=True)
admin_head.font = Font(bold=True)
mod_head.font = Font(bold=True)
desc_head.font = Font(bold=True)
date_head.font = Font(bold=True)
outputsheet.column_dimensions['A'].width = 15
outputsheet.column_dimensions['B'].width = 20
outputsheet.column_dimensions['C'].width = 20
outputsheet.column_dimensions['D'].width = 20
outputsheet.column_dimensions['E'].width = 30
outputsheet.column_dimensions['F'].width = 20

wb.save("services.xlsx")
# print(uicColumn)
# wb.save 
# print(firstsheet.cell(1,uicColumn))
# for column in firstsheet.values:
#     print(uicColumn)
